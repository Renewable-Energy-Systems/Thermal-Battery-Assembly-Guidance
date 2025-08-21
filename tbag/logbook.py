"""
Shared log/timeline helpers + XLSX export.
"""
import json, hashlib, io
import sqlite3
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .db import connect
from .projects_helpers import load as load_project

def _hue(name:str)->int:
    return int(hashlib.md5(name.encode()).hexdigest()[:2],16)*360//255

def overview_rows() -> list[dict]:
    """
    One dict per run, newest first.
    Data comes directly from the `runs` table, not from the events table.
    """
    rows: list[dict] = []
    with connect() as c:
        c.row_factory = sqlite3.Row
        for r in c.execute(
            "SELECT ts_created, project, stack_id, operator, status, "
            "       interrupted_at, session_id "
            "FROM runs ORDER BY ts_created DESC"
        ).fetchall():

            # map DB → fields expected by the template
            kind = "end" if r["status"] == "finished" else "abort"
            rows.append({
                "ts":        r["ts_created"],
                "project":   r["project"],
                "stack_id":  r["stack_id"],
                "operator":  r["operator"],
                "kind":      kind,
                "step":      r["interrupted_at"],
                "session_id":r["session_id"],
                "hue":       _hue(r["project"]),
            })
    return rows

def timeline(sid:str)->List[Dict]:
    with connect() as c:
        cur=c.execute("SELECT ts,event FROM events WHERE event LIKE ? ORDER BY ts",(f"%{sid}%",))
        data=[]
        for ts,raw in cur.fetchall():
            if "::" in raw:
                k,p=raw.split("::",1); p=json.loads(p)
            else: k,p=raw,{}
            data.append({"ts":ts,"kind":k,"payload":p})
        return data

def _sheet(wb:Workbook,title,header,rows):
    ws=wb.create_sheet(title); ws.append(header)
    for r in rows: ws.append(r)
    for col in range(1,len(header)+1):
        ws.column_dimensions[get_column_letter(col)].width=16

def export_overview()->io.BytesIO:
    wb=Workbook(); wb.remove(wb.active)
    dat=overview_rows()
    _sheet(wb,"sessions",
           ["Started","Project","Stack","Operator","Status","Step","Session"],
           [[r["ts"],r["project"],r["stack"],r["operator"],
             r["kind"],(r["step"]+1) if r["step"] else "",r["session"]] for r in dat])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def export_detail(sid: str) -> io.BytesIO:
    """
    Pretty, human‑readable XLSX timeline for a session:
      - Big title row with the session id
      - Columns: Step #, Time, Since Start, Delta (s), Event, Component, Session ID
      - Frozen header, AutoFilter, sensible column widths
      - 'Summary' sheet with start/end, total steps, duration
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # ── 1) Pull events (existing helper) and normalize ──────────────────────────
    tl = timeline(sid)  # [{ "ts": "...", "kind": "...", "payload": {.../str} }, ...]

    rows = []
    for e in tl:
        ts_raw = e.get("ts")
        kind = e.get("kind", "")
        payload = e.get("payload") or {}

        # payload might be a JSON-encoded string — make it a dict safely
        if isinstance(payload, str):
            try:
                payload = json.loads(payload)
            except Exception:
                payload = {}

        comp = payload.get("component")
        sess = payload.get("session_id", sid)

        # Try to parse TS to datetime (for Excel date formatting, deltas, etc)
        ts_dt = None
        if isinstance(ts_raw, str):
            try:
                ts_dt = datetime.fromisoformat(ts_raw)
            except Exception:
                ts_dt = None

        rows.append({"ts_raw": ts_raw, "ts_dt": ts_dt, "kind": kind,
                     "component": comp, "session_id": sess})

    # Empty session safety: return a blank workbook
    wb = Workbook()
    wb.remove(wb.active)  # we'll create sheets ourselves

    ws = wb.create_sheet("Timeline")

    # Title
    cols = ["Step #", "Time", "Since Start (mm:ss)", "Delta (s)", "Event", "Component", "Session ID"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    title_cell = ws.cell(row=1, column=1, value=f"TBAG Session Timeline — {sid}")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="DDDDDD")

    # Header
    header_row = 3
    thin = Side(style="thin")
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=header_row, column=j, value=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # If there are no rows, still return the workbook (with title + header)
    if not rows:
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    # ── 2) Compute step numbers, deltas, since‑start ────────────────────────────
    # Use the first valid datetime as start; if none parse, we won't compute deltas
    ts_dts = [r["ts_dt"] for r in rows if r["ts_dt"] is not None]
    start_dt = ts_dts[0] if ts_dts else None

    def fmt_since_start(d: datetime) -> str:
        if start_dt is None or d is None:
            return ""
        secs = int((d - start_dt).total_seconds())
        return f"{secs//60:02d}:{secs%60:02d}"

    def delta_seconds(prev: datetime, cur: datetime):
        if prev is None or cur is None:
            return None
        return int((cur - prev).total_seconds())

    prev_dt = None
    for i, r in enumerate(rows, start=1):
        time_cell_value = r["ts_dt"] if r["ts_dt"] is not None else r["ts_raw"]
        since_start = fmt_since_start(r["ts_dt"])
        delta_s = delta_seconds(prev_dt, r["ts_dt"]) if i > 1 else None
        prev_dt = r["ts_dt"]

        values = [
            i,                               # Step #
            time_cell_value,                 # Time (datetime if parsed, else str)
            since_start,                     # Since Start
            delta_s,                         # Delta (s)
            r["kind"],                       # Event
            r["component"],                  # Component
            r["session_id"],                 # Session ID
        ]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=header_row + i, column=j, value=v)
            # Formatting
            if j == 2 and isinstance(v, datetime):
                c.number_format = "yyyy-mm-dd hh:mm:ss"
            if j == 4 and isinstance(v, int):
                c.number_format = "0"
            c.border = Border(left=thin, right=thin)

    # Bottom border on last data row
    last_row = header_row + len(rows)
    for j in range(1, len(cols) + 1):
        ws.cell(row=last_row, column=j).border = Border(bottom=thin, left=thin, right=thin)

    # Column widths
    widths = [8, 20, 16, 10, 16, 28, 16]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Freeze panes & AutoFilter
    ws.freeze_panes = ws["A4"]
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(cols))}{last_row}"

    # ── 3) Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"], ws2["B1"] = "Session ID", sid

    # Start/End/Durations only if we had valid datetimes
    if ts_dts:
        start = min(ts_dts); end = max(ts_dts)
        duration_secs = int((end - start).total_seconds())

        ws2["A2"], ws2["B2"] = "Start Time", start
        ws2["A3"], ws2["B3"] = "End Time", end
        ws2["A4"], ws2["B4"] = "Total Steps", len(rows)

        # Excel time is a fraction of a day — store seconds/86400 and format as [m]:ss
        ws2["A5"], ws2["B5"] = "Duration (mm:ss)", duration_secs / 86400.0
        ws2["B2"].number_format = ws2["B3"].number_format = "yyyy-mm-dd hh:mm:ss"
        ws2["B5"].number_format = "[m]:ss"
    else:
        ws2["A2"], ws2["B2"] = "Note", "Timestamps were not ISO-8601; summary timing unavailable."
        ws2["A4"], ws2["B4"] = "Total Steps", len(rows)

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 32
    for r in range(1, 6):
        ws2[f"A{r}"].font = Font(bold=True)

    # ── 4) Return as BytesIO ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

