#!/usr/bin/env python3
"""
TBAG · Thermal-Battery Assembly Guidance
────────────────────────────────────────────────────────────────────────────
Supervisor panel  →  /admin  (+ /admin/sessions)
Kiosk (Pi)        →  /       polls /api/claim → /api/progress
Projects CRUD     →  /projects
Logs              →  /logs    (+ Excel exports)
"""

# ── std-lib ───────────────────────────────────────────────────────────────
import os, sys, json, uuid, datetime, sqlite3, pathlib, hashlib, io
from typing import List, Dict, Optional

# ── 3rd-party ─────────────────────────────────────────────────────────────
from flask import (
    Flask, render_template, render_template_string,
    request, redirect, session, abort,
    jsonify, send_from_directory, send_file
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import werkzeug.datastructures

# ── GPIO or dummy stubs ───────────────────────────────────────────────────
try:
    from gpiozero import LED, Button, Device
    if not (sys.platform.startswith("linux") and os.path.exists("/proc/cpuinfo")):
        from gpiozero.pins.mock import MockFactory
        Device.pin_factory = MockFactory()
except ImportError:                    # dev on non-Pi hosts
    class _Dummy:
        def __getattr__(self, *_): return lambda *a, **k: None
        is_active = False
    LED = Button = _Dummy              # type: ignore

red_led, pedal = LED(17), Button(27)   # change BCM pins if necessary

# ── paths & Flask init ────────────────────────────────────────────────────
BASE      = pathlib.Path(__file__).parent
PROJECTS  = BASE / "projects";  PROJECTS.mkdir(exist_ok=True)
DB_FILE   = BASE / "events.db"
DEVICE_ID = os.environ.get("TBAG_DEVICE", "glovebox-pi")

app = Flask(__name__, template_folder=str(BASE / "templates"))
app.secret_key = "change-me"           # ⚠ set real secret in prod
app.config.update(TEMPLATES_AUTO_RELOAD=True, SEND_FILE_MAX_AGE_DEFAULT=0)
app.jinja_env.auto_reload = True

# ── DB bootstrap ──────────────────────────────────────────────────────────
def init_db() -> None:
    with sqlite3.connect(DB_FILE) as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS events(
          ts    TEXT,
          event TEXT
        );
        CREATE TABLE IF NOT EXISTS runs(
          session_id     TEXT PRIMARY KEY,
          project        TEXT,
          stack_id       TEXT,
          operator       TEXT,
          ts_created     TEXT,
          ts_started     TEXT,
          ts_finished    TEXT,
          status         TEXT CHECK(status IN
                        ('pending','active','finished','aborted')),
          interrupted_at INTEGER,
          device         TEXT
        );
        """)
def log(e: str) -> None:
    with sqlite3.connect(DB_FILE) as c:
        c.execute("INSERT INTO events VALUES(?,?)",
                  (datetime.datetime.now().isoformat(timespec="seconds"), e))
init_db()

# ── helpers: projects ─────────────────────────────────────────────────────
def projects_list() -> List[pathlib.Path]:
    return sorted((d for d in PROJECTS.iterdir() if d.is_dir()),
                  key=lambda p: p.name.lower())

def load_config(pid: str) -> Optional[Dict]:
    cfg = PROJECTS / pid / "config.json"
    return json.load(cfg.open()) if cfg.exists() else None

def save_config(pid: str, data: Dict) -> None:
    d = PROJECTS / pid
    (d / "images").mkdir(parents=True, exist_ok=True)
    with (d / "config.json").open("w") as f:
        json.dump(data, f, indent=2)

# ──────────────────────────────────────────────────────────────────────────
#  K I O S K   A P I
# ──────────────────────────────────────────────────────────────────────────
@app.route("/")
def index():                              # single kiosk page
    return render_template("index.html")

@app.post("/api/claim")
def api_claim():
    """
    Atomically change first ‘pending’ run to ‘active’
    and hand it to the kiosk.
    """
    with sqlite3.connect(DB_FILE) as c:
        c.row_factory = sqlite3.Row
        cur = c.execute("""
            UPDATE runs
               SET status='active',
                   ts_started=?,
                   device=?
             WHERE session_id = (
                SELECT session_id FROM runs
                 WHERE status='pending'
                 ORDER BY ts_created
                 LIMIT 1
             )
        RETURNING *""",
            (datetime.datetime.now().isoformat(timespec="seconds"), DEVICE_ID)
        )
        row = cur.fetchone()             # ← fixed

    if row is None:
        return jsonify(status="none")

    cfg = load_config(row["project"]) or {"sequence": []}
    return jsonify(status="claimed",
                   session=dict(row),
                   sequence=cfg["sequence"])

@app.post("/api/progress")
def api_progress():
    """
    Body: {session_id, action:next|finish|abort, step?}
    """
    data = request.json
    sid  = data["session_id"]
    act  = data["action"]

    if act == "next":
        red_led.toggle()
        log(f"next_pressed::{sid}")
        return jsonify(status="ok")

    with sqlite3.connect(DB_FILE) as c:
        if act == "finish":
            c.execute("""UPDATE runs
                            SET status='finished',
                                ts_finished=?
                          WHERE session_id=?""",
                      (datetime.datetime.now().isoformat(timespec="seconds"), sid))
            log(f"session_end::{json.dumps({'session_id': sid})}")

        elif act == "abort":
            c.execute("""UPDATE runs
                            SET status='aborted',
                                ts_finished=?,
                                interrupted_at=?
                          WHERE session_id=?""",
                      (datetime.datetime.now().isoformat(timespec="seconds"),
                       data.get("step"), sid))
            log(f"session_abort::{json.dumps({'session_id': sid,'step': data.get('step')})}")
    return jsonify(status="ok")

@app.get("/pedal")                        # simple polling helper
def pedal_status():
    return jsonify(pressed=bool(getattr(pedal, "is_active", False)))

# ──────────────────────────────────────────────────────────────────────────
#  S U P E R V I S O R   P A N E L
# ──────────────────────────────────────────────────────────────────────────
@app.get("/admin")                #   /admin   (no slash)
@app.get("/admin/")               #   /admin/  (with slash)
def admin_dash():
    return render_template("admin_dashboard.html")

@app.get("/admin/sessions")
def admin_sessions():
    with sqlite3.connect(DB_FILE) as c:
        c.row_factory = sqlite3.Row
        rows = c.execute("SELECT * FROM runs ORDER BY ts_created DESC").fetchall()
    projs = [{"id": p.name, "name": load_config(p.name)["name"]}
             for p in projects_list()]
    return render_template("admin_sessions.html", rows=rows, projects=projs)

@app.post("/admin/sessions/new")
def admin_sessions_new():
    sid = uuid.uuid4().hex[:8]
    payload = dict(
        session_id=sid,
        project   =request.form["project"],
        stack_id  =request.form["stack_id"],
        operator  =request.form["operator"],
        ts_created=datetime.datetime.now().isoformat(timespec="seconds"),
        status    ="pending"
    )
    with sqlite3.connect(DB_FILE) as c:
        c.execute("""INSERT INTO runs(session_id,project,stack_id,operator,
                     ts_created,status) VALUES(:session_id,:project,:stack_id,
                     :operator,:ts_created,:status)""", payload)
    return redirect("/admin/sessions")

@app.post("/admin/sessions/<sid>/delete")
def admin_sessions_delete(sid):
    with sqlite3.connect(DB_FILE) as c:
        deleted = c.execute(
            "DELETE FROM runs WHERE session_id=? AND status='pending'", (sid,)
        ).rowcount
    if deleted == 0:
        abort(400, "Cannot delete — session active or finished.")
    return redirect("/admin/sessions")

# ──────────────────────────────────────────────────────────────────────────
#  P R O J E C T S  (unchanged CRUD)
# ──────────────────────────────────────────────────────────────────────────
@app.get("/projects")
def list_projects():
    projs = [{"id": p.name, "name": load_config(p.name)["name"]}
             for p in projects_list()]
    return render_template("project_list.html", projects=projs)

@app.route("/projects/new", methods=["GET", "POST"])
def new_project():
    if request.method == "POST":
        name = request.form["proj_name"].strip()
        if any(load_config(p.name)["name"].lower() == name.lower()
               for p in projects_list()):
            return render_template("project_new.html",
                                   error=f'Project “{name}” already exists.')
        pid = name.replace(" ", "_").lower()[:40] or uuid.uuid4().hex[:6]
        save_config(pid, {"name": name, "sequence": []})
        return redirect(f"/projects/{pid}/edit")
    return render_template("project_new.html", error=None)

@app.route("/projects/<pid>/edit", methods=["GET", "POST"])
def edit_project(pid):
    cfg = load_config(pid) or {"name": pid, "sequence": []}
    if request.method == "POST":
        seq, idx = [], 0
        while True:
            lbl = request.form.get(f"label_{idx}")
            if lbl is None:
                break
            img = cfg["sequence"][idx]["img"] if idx < len(cfg["sequence"]) else None
            fs: werkzeug.datastructures.FileStorage = request.files.get(f"img_{idx}")  # type: ignore
            if fs and fs.filename:
                img = f"step{idx}_{uuid.uuid4().hex[:6]}{pathlib.Path(fs.filename).suffix}"
                fs.save(PROJECTS / pid / "images" / img)
            seq.append({"label": lbl, "img": img})
            idx += 1
        cfg["sequence"] = seq
        save_config(pid, cfg)
        return redirect("/projects")
    return render_template("project_edit.html", project=cfg, sequence=cfg["sequence"])

@app.get("/proj_assets/<pid>/<path:fname>")
def proj_assets(pid, fname):
    return send_from_directory(PROJECTS / pid / "images", fname)

# ──────────────────────────────────────────────────────────────────────────
#  L O G S  +  E X P O R T S
# ──────────────────────────────────────────────────────────────────────────
def _hue_for_proj(name:str)->int:
    return int(hashlib.md5(name.encode()).hexdigest()[:2],16)*360//255

def logs_overview_data()->List[Dict]:
    rows=[]
    with sqlite3.connect(DB_FILE) as c:
        for ts,ev in c.execute(
            "SELECT ts,event FROM events WHERE event LIKE 'session_%' ORDER BY ts DESC"):
            kind,payload=ev.split("::",1)
            if kind=="session_start": continue
            d=json.loads(payload)
            rows.append({"ts":ts,"kind":kind.replace("session_",""),
                         "project":d.get("project"),"stack_id":d.get("stack_id"),
                         "operator":d.get("operator"),"session_id":d["session_id"],
                         "step":d.get("interrupted_at"),"hue":_hue_for_proj(d.get("project",""))})
    return rows

def log_detail_data(sid:str)->List[Dict]:
    evs=[]
    with sqlite3.connect(DB_FILE) as c:
        for ts,raw in c.execute("SELECT ts,event FROM events WHERE event LIKE ? ORDER BY ts",(f"%{sid}%",)):
            if "::" in raw:
                k,p=raw.split("::",1); p=json.loads(p)
            else: k,p=raw,{}
            evs.append({"ts":ts,"kind":k,"payload":p})
    return evs

@app.get("/logs")
def logs_overview():
    return render_template("logs_overview.html", rows=logs_overview_data())

@app.get("/logs/<sid>")
def log_detail(sid):
    evs = log_detail_data(sid)
    if not evs: return f"No logs for {sid}", 404
    return render_template("log_detail.html", events=evs, sid=sid)

# --- Excel ----------------------------------------------------------------
def _sheet(wb:Workbook,title:str,header:list,rows:list):
    ws=wb.create_sheet(title); ws.append(header)
    for r in rows: ws.append(r)
    for col in range(1,len(header)+1):
        ws.column_dimensions[get_column_letter(col)].width=16

@app.get("/logs/export")
def export_overview():
    wb=Workbook(); wb.remove(wb.active)
    data=logs_overview_data()
    _sheet(wb,"sessions",
           ["Started","Project","Stack","Operator","Status","Step","Session ID"],
           [[r["ts"],r["project"],r["stack_id"],r["operator"],
             r["kind"],(r["step"]+1) if r["step"] is not None else "",r["session_id"]]
             for r in data])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name="sessions.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/logs/<sid>/export")
def export_detail(sid):
    evs=log_detail_data(sid)
    if not evs: return redirect("/logs")
    wb=Workbook(); wb.remove(wb.active)
    _sheet(wb,"timeline",
           ["Time","Event","Payload"],
           [[e["ts"],e["kind"],json.dumps(e["payload"],separators=(',',':'))] for e in evs])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,as_attachment=True,download_name=f"{sid}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ──────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":       # dev; production uses gunicorn
    app.run(host="0.0.0.0", port=8000, debug=False)
